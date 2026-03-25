import openpyxl
from openpyxl.drawing.image import Image as XLImage
import os
import sys
import base64
from datetime import datetime
import win32com.client 
import pythoncom

# --- CONFIGURATION ---
CELL_MAPPING = {
    # --- PATIENT INFO ---
    "PATIENT": "D3",        
    "AGE": "D4",            
    "PATIENT_ID": "D6",     
    "DATE": "H5",
    "GENDER" : "D5",
    "REF_BY": "I6",

    # --- Date Section ---
    "REG_DATE": "I3",       # Registration Date
    "SAMPLE_DATE": "I4",    # Sample Tested Date
    "REPORT_DATE": "I5",    # Report Printed Date
    "SAMPLE_TYPE": "D7",    # Specimen Type

    # --- RBC SECTION ---
    "RBC": "D14",
    "HGB": "D13",
    "HCT": "D15",
    "MCV": "D16",
    "MCH": "D17",
    "MCHC": "D18",
    "RDW-CV": "D19",
    "RDW-SD": "D20",

    # --- WBC SECTION ---
    "WBC": "D25",
    "NEU%": "D26",
    "LYM%": "D27",
    "MON%": "D28",
    "EOS%": "D29",
    "BAS%": "D30",

    # --- ABSOLUTE COUNTS ---
    "NEU#": "M68",
    "LYM#": "M36",
    "MON#": "M37",
    "EOS#": "M38",
    "BAS#": "M39",

    # --- PLATELETS ---
    "PLT": "D43",
    "MPV": "D44",
    "PDW-CV": "D45",
    "PDW-SD": "D46",
    "PCT": "D47",
    "P-LCR": "D48",
    "P-LCC": "D49"
}
IMAGES_MAPPING = {
    "WBC Histogram. BMP" : "I25",
    "RBC Histogram. BMP" : "I13",
    "PLT Histogram. BMP" : "I43" 
}


class ExcelReport:
    def __init__(self, template_path="report_template.xlsx", output_path="current_report.xlsx"):
        self.template_path = template_path
        self.output_path = output_path
        self.pdf_folder = os.path.join(os.getcwd(), "PDF_Reports")
        self.tempImgFolder = os.path.join(os.getcwd(), "temp_images")
        if not os.path.exists(self.tempImgFolder):
            os.makedirs(self.tempImgFolder)

    def generate(self, data_dict):
        if not os.path.exists(self.template_path):
            self._create_dummy_template()
            print(f"WARNING: Template '{self.template_path}' not found. Created a dummy one.")

        try:
            wb = openpyxl.load_workbook(self.template_path)
            ws = wb.active 
        except Exception as e:
            print(f"Error loading Excel template: {e}")
            return False

        flat_data = self._flatten_data(data_dict)

        for key, cell_address in CELL_MAPPING.items():
            if key in flat_data:
                val = flat_data[key]
                # try:
                #     clean_val = float(val)
                # except ValueError:
                clean_val = val

                if isinstance(val, (dict, list)):
                    # If somehow a dict got here, force it to string to prevent crash
                    clean_val = str(val) 
                else:
                    try:
                        clean_val = float(val)
                    except (ValueError, TypeError):
                        clean_val = val


                # Handle Merged Cells / Tuple Error
                target = ws[cell_address]
                if isinstance(target, tuple):
                    target[0][0].value = clean_val
                else:
                    target.value = clean_val
            else:
                pass
        imagesDict = data_dict.get("IMAGES", {})
        for imgName , base64Str in imagesDict.items():
            #Check if we have a mapping for this image
            #Note : The raw Data migt have extra spaces. so strip keys
            targetCell = IMAGES_MAPPING.get(imgName.strip())

            if targetCell and base64Str:
                try:
                    imgData = base64.b64decode(base64Str)

                    #save to temp file (openpyxl needs a file usually)
                    #we use a save filename based on te test name
                    safeName = "".join([c for c in imgName if c.isalnum()]) + ".png"
                    imgPath = os.path.join(self.tempImgFolder, safeName)
                    
                    with open(imgPath, "wb") as f:
                        f.write(imgData)

                    #create openPyXL Image
                    img = XLImage(imgPath)

                    #Optional: Resize image if needed
                    #img.width = 300
                    #img.height = 200
                    #add to worksheet
                    ws.add_image(img, targetCell)
                    print(f"inserted Image: {imgName} at {targetCell}")


                except Exception as e:
                    print(f"Failed to insert image {imgName}: {e}")


        try:
            wb.save(self.output_path)
            return True
        except PermissionError:
            print("ERROR: Could not save file. Is it open in Excel?")
            return False

    def save_as_pdf(self, data_dict):
        if not self.generate(data_dict):
            return False, "Could not save intermediate Excel file."

        if not os.path.exists(self.pdf_folder):
            try:
                os.makedirs(self.pdf_folder)
            except OSError as e:
                return False, f"Could not create PDF folder: {e}"

        p_name = str(data_dict.get("PATIENT", "Unknown")).strip()
        p_name = "".join([c for c in p_name if c.isalpha() or c.isdigit() or c==' ']).rstrip()
        p_id = str(data_dict.get("PATIENT_ID", "NoID"))
        date_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        
        pdf_name = f"{p_id}_{p_name}_{date_str}.pdf"
        pdf_path = os.path.join(self.pdf_folder, pdf_name)

        abs_excel_path = os.path.abspath(self.output_path)
        abs_pdf_path = os.path.abspath(pdf_path)

        excel_app = None
        wb = None
        
        try:
            pythoncom.CoInitialize() 
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            
            wb = excel_app.Workbooks.Open(abs_excel_path)
            ws = wb.ActiveSheet
            
            try: ws.Unprotect()
            except: pass

            for shape in ws.Shapes:
                # 1 = msoPictue
                #we simply for Placement = 1 (move and size) or PrintObject = True if accessible
                #Some objects might throw errors, so we wrap in try/except
                try:
                    shape.Placement = 1 #xlMoveandSize
                    # .PrintObject is not always directly exposed in all COM interface for 'Shapes'
                except:
                    pass

                excel_app.ScreenUpdating = True

                wb.Save()
            wb.ExportAsFixedFormat(0, abs_pdf_path, 0, True, False)
            return True, abs_pdf_path

        except Exception as e:
            return False, str(e)
            
        finally:
            if wb:
                try: wb.Close(False)
                except: pass
            if excel_app:
                try: excel_app.Quit()
                except: pass
            pythoncom.CoUninitialize()

    def print_file(self, data_dict):
        """Generates Excel and sends to Default Printer via COM."""
        # 1. Update the temp excel file
        if not self.generate(data_dict):
            return False, "Could not save intermediate Excel file."

        abs_excel_path = os.path.abspath(self.output_path)
        excel_app = None
        wb = None

        try:
            pythoncom.CoInitialize()
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False

            wb = excel_app.Workbooks.Open(abs_excel_path)
            
            ws = wb.ActiveSheet
            # Force shapes to print
            for shape in ws.Shapes:
                try: shape.Placement = 1
                except: pass
            
            wb.Save() # Save changes to shapes before print
            
            # PrintOut() uses default printer settings
            wb.PrintOut()
            
            return True, "Sent to Printer"

        except Exception as e:
            return False, str(e)

        finally:
            if wb:
                try: wb.Close(False)
                except: pass
            if excel_app:
                try: excel_app.Quit()
                except: pass
            pythoncom.CoUninitialize()

    def open_file(self):
        if os.path.exists(self.output_path):
            if sys.platform == "win32":
                os.startfile(self.output_path)
            else:
                import subprocess
                opener = "open" if sys.platform == "darwin" else "xdg-open"
                subprocess.call([opener, self.output_path])

    def open_pdf_folder(self):
        if os.path.exists(self.pdf_folder):
            if sys.platform == "win32":
                os.startfile(self.pdf_folder)

    def _flatten_data(self, nested_data):
        flat = {}
        flat["PATIENT"] = nested_data.get("PATIENT", "")
        flat["AGE"] = nested_data.get("AGE", "")
        flat["GENDER"] = nested_data.get("GENDER", "") 
        flat["PATIENT_ID"] = nested_data.get("PATIENT_ID", "")
        flat["REF_BY"] = nested_data.get("REF_BY", "")
        # flat["DATE"] = datetime.now().strftime("%Y-%m-%d")

        # --- NEW DATE FIELDS ---
        flat["REG_DATE"] = nested_data.get("REG_DATE", "")
        flat["SAMPLE_DATE"] = nested_data.get("SAMPLE_DATE", "")
        flat["REPORT_DATE"] = nested_data.get("REPORT_DATE", "")
        flat["SAMPLE_TYPE"] = nested_data.get("SAMPLE_TYPE", "")



        for section_key, section_val in nested_data.items():
            if isinstance(section_val, dict):
                for test_name, test_data in section_val.items():
                    if isinstance(test_data, dict):
                        flat[test_name] = test_data.get("value", "")
        return flat

    def _create_dummy_template(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CBC Report"
        ws['A1'] = "WARNING: THIS IS A GENERATED DUMMY TEMPLATE."
        ws['A2'] = "Please replace 'report_template.xlsx' with your own design."
        for key, cell in CELL_MAPPING.items():
            ws[cell] = f"{{ {key} }}" 
        wb.save(self.template_path)